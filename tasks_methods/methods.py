import csv
import json
import os
import re
from urllib.error import URLError
import urllib.request
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import Workbook
from robocorp import workitems
from helpers.article import Article
from helpers.payload import Payload
from helpers.selector import Selector
from webdriver_util.webdrv_util import *
from urllib.parse import unquote
from dotenv import load_dotenv
from RPA.HTTP import HTTP

load_dotenv("config/.env")


class ProducerMethods:
    @staticmethod
    def read_csv_create_work_item(debug: bool = False):
        """
        Reads a CSV file and creates work items from its data.

        Args:
            debug (bool): If True, the function returns the payload instead of creating work items.

        Returns:
            Payload | None: Returns payload if debug is True, otherwise None.
        """
        csv_file_path = os.path.join("devdata", "csv_input.csv")
        if os.path.exists(csv_file_path):
            try:
                with open(csv_file_path, mode="r", newline="") as file:
                    reader = csv.reader(file)
                    header = next(reader)
                    for row in reader:
                        payload = Payload(
                            phrase_test=row[0],
                            section=row[1],
                            sort_by=int(row[2]),
                            results=int(row[3]),
                        )
                        if not debug:
                            return workitems.outputs.create(
                                payload={
                                    "phrase_test": payload.phrase_test,
                                    "section": payload.section,
                                    "sort_by": payload.sort_by,
                                    "results": payload.results,
                                }
                            )
                        else:
                            return payload
            except FileNotFoundError:
                logger.critical(f"The CSV file: {csv_file_path} was not found.")
                return None
            except ValueError as e:
                logger.critical(f"ValueError: {e}")
                return None
            except csv.Error as e:
                logger.critical(f"csv.Error: {e}")
                return None
            except Exception as e:
                logger.critical(f"Unexpected error: {e}")
                return None
        else:
            logger.critical(f"The CSV file: {csv_file_path} was not found.")
            return None


class ScraperMethods:
    @staticmethod
    def get_work_item() -> Payload | None:
        """
        Retrieves the current work item and converts it to a Payload object.

        Returns:
            Payload | None: Returns the payload if a work item exists, otherwise None.
        """
        try:
            item = workitems.inputs.current
            if item:
                logger.info(f"Received payload:{item.payload}")
                pay = Payload(
                    phrase_test=item.payload["phrase_test"],
                    section=item.payload["section"],
                    sort_by=int(item.payload["sort_by"]),
                    results=int(item.payload["results"]),
                )
                return pay
            else:
                logger.critical("An error occurred during the process!")
        except KeyError as e:
            logger.critical(f"KeyError: {e}")
            return None
        except TypeError as e:
            logger.critical(f"TypeError: {e}")
            return None
        except Exception as e:
            logger.critical(f"Unexpected error: {e}")
            return None

    @staticmethod
    def inicial_search(driver: Selenium, phrase: str):
        """
        Performs the initial search on the website.

        Args:
            driver (Selenium): The Selenium driver instance.
            phrase (str): The search phrase.

        Returns:
            bool: True if the search was initiated successfully, otherwise False.
        """
        try:
            logger.info("Starting Scraper")
            search = find_element(
                driver.driver, Selector(css='button[data-element="search-button"]')
            )
            if search:
                center_element(driver.driver, search)
                click_elm(driver.driver, search)
                search_field = find_element(
                    driver.driver,
                    Selector(css="input[data-element='search-form-input']"),
                )
                if search_field:
                    center_element(driver.driver, search_field)
                    slow_send_keys(search_field, phrase + Keys.ENTER, False)
                    return True
        except AttributeError as e:
            logger.critical(f"AttributeError: {e}")
            return None
        except TypeError as e:
            logger.critical(f"TypeError: {e}")
            return None
        except Exception as e:
            logger.critical(f"Unexpected error: {e}")
            return None
        return False

    @staticmethod
    def fine_search(
        driver: WebDriver,
        section: str,
        sort_by: int = 0,
    ):
        """
        Performs fine-tuned search with additional filters.

        Args:
            driver (WebDriver): The WebDriver instance.
            phrase (str): The search phrase.
            section (str): The section to filter.
            sort_by (int): The sort option (default is 0).

        Returns:
            bool indicating success.
        """
        try:
            no_results_match = find_element(
                driver.driver,
                Selector(css="div[class='search-results-module-no-results']"),
            )
            if no_results_match:
                logger.critical("No search match found.")
                return False

            # Expand Filter
            label_search = find_element(
                driver.driver, Selector(css="span[class='see-all-text']")
            )
            if label_search:
                center_element(driver.driver, label_search)
                click_elm(driver.driver, label_search)
                wait_for_modal(driver.driver)
                if len(section.strip()) > 0:
                    list_topics = extract_names_from_list_items(driver)
                    if list_topics:
                        element_topic, topic = search_and_click_topics(
                            driver.driver, list_topics, section
                        )
                        if not element_topic and not topic:
                            return False, 0

                if sort_by > 0:  # not Relevance (default)
                    select_sort_by = find_element(
                        driver.driver, Selector(css="select[name='s']")
                    )
                    if select_sort_by:
                        if sort_by in [1, 2]:
                            center_element(driver.driver, select_sort_by)
                            select_option_value(select_sort_by, sort_by)
                        else:
                            logger.error(f"Sort parameter does not exist: {sort_by}")
                            logger.info("Relevance is selected")
                return (True,)
        except AttributeError as e:
            logger.critical(f"AttributeError: {e}")
            return None
        except TypeError as e:
            logger.critical(f"TypeError: {e}")
            return None
        except Exception as e:
            logger.critical(f"Unexpected error: {e}")
        return False

    @staticmethod
    def collect_articles(driver: WebDriver, results: int = 0) -> list[Article] | None:
        """
        Collects articles from the search results.

        Args:
            driver (WebDriver): The WebDriver instance.
            data_range (int): The data range option.

        Returns:
            list[Article] | None: List of collected articles or None if an error occurs.
        """
        try:
            list_articles = []
            more_results = True
            cont = 1
            while more_results:
                logger.info("Search results found")
                search_results_section = find_element(
                    driver.driver,
                    Selector(css="div[class='search-results-module-results-header']"),
                )
                if search_results_section:
                    logger.info("Search results found")
                    wait_for_modal(driver.driver)
                    li_search_results = find_all_css(
                        driver.driver,
                        'ul[class*="search-results-module-results-menu"] li',
                    )
                    if li_search_results:
                        for li in li_search_results:
                            logger.info(f"Creating an article object: {cont}")
                            article = Article()
                            try:
                                title = li.find_element(
                                    By.CSS_SELECTOR, "h3[class='promo-title']"
                                )
                                time = li.find_element(
                                    By.CSS_SELECTOR, "p[class^='promo-timestamp']"
                                )
                                description = li.find_element(
                                    By.CSS_SELECTOR, "p[class='promo-description']"
                                )
                            except AttributeError as e:
                                logger.critical(f"AttributeError: {e}")
                                return None
                            except TypeError as e:
                                logger.critical(f"TypeError: {e}")
                                return None
                            except Exception as e:
                                logger.critical(f"Unexpected error: {e}")
                                return None

                            try:
                                center_element(driver.driver, li)
                                photo = find_elm_picture(
                                    li, Selector(css='img[src*=".jpg"]')
                                )
                                if not photo is None:
                                    article.picture_filename = photo
                                    logger.info(
                                        f"Picture found: {article.picture_filename}"
                                    )
                            except AttributeError as e:
                                logger.critical(f"AttributeError: {e}")
                            except TypeError as e:
                                logger.critical(f"TypeError: {e}")
                            except Exception as e:
                                logger.critical(f"Unexpected error: {e}")
                                logger.info("Picture information in article not found.")

                            logger.info("Article information found.")
                            article.title = title.text.strip()
                            article.description = description.text.strip()
                            time_str = time.text.strip()
                            parse = parse_time_ago(time_str)
                            if not parse == None:
                                article.date = parse
                            else:
                                article.date = datetime.strptime(
                                    time.text.strip(), "%B %d, %Y"
                                )
                            logger.info(
                                f"Title: {article.title} -- Date: {article.date}"
                            )
                            list_articles.append(article)
                            if results == cont:
                                more_results = False
                                break
                            cont += 1
                        button_next = find_element(
                            driver.driver,
                            Selector(
                                css='div[class="search-results-module-next-page"]'
                            ),
                        )
                        if button_next:
                            center_element(driver.driver, button_next)
                            click_elm(driver.driver, button_next)
                            if results < cont:
                                more_results = False
            return list_articles
        except AttributeError as e:
            logger.critical(f"AttributeError: {e}")
            return None
        except TypeError as e:
            logger.critical(f"TypeError: {e}")
            return None
        except Exception as e:
            logger.critical(f"Unexpected error: {e}")
            return None


class ExcelOtherMethods:
    @staticmethod
    def __contains_money(text):
        """
        Checks if the text contains monetary amounts.

        Args:
            text (str): The input text.

        Returns:
            bool: True if monetary amounts are found, otherwise False.
        """
        try:
            pattern = r"\$[0-9,.]+|\b\d+\s*(?:dollars|USD)\b"
            matches = re.findall(pattern, text)
            return bool(matches)
        except re.error as e:
            logger.critical(f"Regex error: {e}")
            return False

    @staticmethod
    def __download_image(url: str, article_title: str) -> str:
        """
        Downloads an image from a given URL with a filename based on the article title.

        Args:
            url (str): The URL of the image.
            article_title (str): The title of the article.

        Returns:
            str: The local path where the image is saved.
        """
        try:
            http = HTTP()
            # Extract the file extension from the URL
            file_extension = os.path.splitext(url)[-1]

            # Sanitize the article title and ensure it is no more than 20 characters
            sanitized_title = re.sub(r'[<>:"/\\|?*]', "", article_title)
            short_title = sanitized_title[:20]

            # Create the full filename
            filename = f"{short_title}{file_extension}"

            # Define the download path
            project_dir = str(os.getcwd())
            full_path = Path(project_dir, "output", "downloads")

            if not os.path.isdir(full_path):
                os.makedirs(full_path)

            if os.path.isdir(full_path):
                full_path = os.path.join(full_path, filename)
                logger.info(f"Downloading image: {url}")
                http.download(
                    url=url,
                    target_file=f"{full_path}",
                    overwrite=True,
                )
                return full_path
        except URLError as e:
            logger.critical(f"URLError: {e}")
            return None
        except Exception as e:
            logger.critical(f"Unexpected error: {e}")
            return None

    @staticmethod
    def prepare_articles(list_articles: list[Article], phrase: str) -> list[Article]:
        """
        Prepares articles by adding additional metadata.

        Args:
            list_articles (list[Article]): List of articles to be prepared.
            phrase (str): The phrase to count in titles and descriptions.

        Returns:
            list[Article]: List of prepared articles.
        """
        try:
            new_list_articles = []
            if list_articles:
                for article in list_articles:
                    art = Article()
                    art.title = article.title
                    art.date = article.date
                    art.title_count_phrase = len(
                        re.findall(
                            re.escape(phrase), article.title.strip(), re.IGNORECASE
                        )
                    )
                    art.description = article.description
                    art.description_count_phrase = len(
                        re.findall(
                            re.escape(phrase),
                            article.description.strip(),
                            re.IGNORECASE,
                        )
                    )
                    art.find_money_title_description = (
                        ExcelOtherMethods.__contains_money(article.title)
                    )
                    if len(article.picture_filename) > 0:
                        art.picture_filename = article.picture_filename
                        art.picture_local_path = ExcelOtherMethods.__download_image(
                            art.picture_filename, article.title.strip()
                        )
                    new_list_articles.append(art)
                    logger.info(f"Article created: {art.to_dict()}")
                return new_list_articles

        except AttributeError as e:
            logger.critical(f"AttributeError: {e}")
            return None
        except TypeError as e:
            logger.critical(f"TypeError: {e}")
            return None
        except re.error as e:
            logger.critical(f"Regex error: {e}")
            return None
        except Exception as e:
            logger.critical(f"Unexpected error: {e}")
            return None

    @staticmethod
    def export_excel(list_articles: list[Article]):
        """
        Exports the list of articles to an Excel file.

        Args:
            list_articles (list[Article]): List of articles to be exported.
        """
        try:
            project_dir = str(os.getcwd())
            full_path = Path(project_dir, "output")
            excel_file_path = os.path.join(full_path, "Articles.xlsx")
            wb = Workbook()
            ws = wb.active
            str_data = Article.articles_to_json(list_articles)
            data = json.loads(str_data)
            headers = list(data[0].keys()) if data else []
            for col_num, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_num, value=header)
            for row_index, row_data in enumerate(data, start=2):
                for col_index, header in enumerate(headers, start=1):
                    ws.cell(
                        row=row_index, column=col_index, value=row_data.get(header, "")
                    )
            logger.info("Excel file created.")
            logger.info("Creating Output...")
            wb.save(excel_file_path)
        except json.JSONDecodeError as e:
            logger.critical(f"JSONDecodeError: {e}")
            return None
        except FileNotFoundError as e:
            logger.critical(f"FileNotFoundError: {e}")
            return None
        except PermissionError as e:
            logger.critical(f"PermissionError: {e}")
            return None
        except Exception as e:
            logger.critical(f"Unexpected error: {e}")
            return None
